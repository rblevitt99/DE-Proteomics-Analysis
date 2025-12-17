import pandas as pd
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import xlsxwriter


# This first function will take the Comparisons Proteomics data set and find all the differentially expressed (DE)
# proteins. The DE proteins will be found by filtering for any protein that has either a log2 fold change 
# > 1 or <-1 and has a p-value of < 0.05.
input_dfs=["Ca_vs_lowCa.xlsx","Ca_vs_Low_Fe.xlsx", "LowFe_vs_Fe.xlsx","Fe_vs_Mix.xlsx","Mix_vs_Ni.xlsx",
"Ni_vs_Nd.xlsx","Nd_vs_W.xlsx","W_vs_Cu_.xlsx","Cu_vs_lowCa.xlsx"]

outsheet=["LowCa_vs_Ca","Ca_vs_Low_Fe", "LowFe_vs_Fe","Fe_vs_Mix","Mix_vs_Ni",
"Ni_vs_Nd","Nd_vs_W","W_vs_Cu","lowCa_vs_Cu"]



workbook=openpyxl.Workbook()
workbook.save(filename="Filtered_DE_Prot_data.xlsx")



def DE_Conductor(input_df,output_sheet):
    df=pd.read_excel(input_df,sheet_name="Sheet2")
    filtered_data = df[((df['Averaged Log2Ratio'] > 1) | (df['Averaged Log2Ratio'] < -1)) & (df['p-value'] < 0.05)] # This code filters through the excel file and only picks out proteins that have a log2 fold change above 1 or below -1 and have a p-value that is less than 0.05
    with pd.ExcelWriter( "Filtered_DE_Prot_data.xlsx", mode="a", engine="openpyxl", if_sheet_exists="new") as writer:
        filtered_data.to_excel(writer, sheet_name=output_sheet, index=False)

for input_df,output_sheet in zip(input_dfs,outsheet):
    DE_Conductor(input_df,output_sheet)

for sheetname in outsheet:
    print(len(pd.read_excel("Filtered_DE_Prot_data.xlsx",sheet_name=sheetname)))

#This next function will map the COG ID's to the DE Protein data set that was just created.

input_sheet=["LowCa_vs_Ca","Ca_vs_Low_Fe", "LowFe_vs_Fe","Fe_vs_Mix","Mix_vs_Ni",
"Ni_vs_Nd","Nd_vs_W","W_vs_Cu","lowCa_vs_Cu"]


out_sheet=["LowCa_vs_Ca","Ca_vs_Low_Fe", "LowFe_vs_Fe","Fe_vs_Mix","Mix_vs_Ni",
"Ni_vs_Nd","Nd_vs_W","W_vs_Cu","lowCa_vs_Cu"]

def COG_Mapper(in_sheet,outsheet):
    deg_file = "Filtered_DE_Prot_data.xlsx"  
    cog_file = "COG_ID's.xlsx"
    with pd.ExcelWriter(deg_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        deg_df = pd.read_excel(deg_file, sheet_name=in_sheet)
        cog_df = pd.read_excel(cog_file)
        merged_df = pd.merge(deg_df, cog_df, on="refseq_protein_id", how="left")
        merged_df["COG_category"].fillna("Unknown", inplace=True)
        merged_df.to_excel(writer, sheet_name=outsheet, index=False)
for in_sheet,outsheet in zip(input_sheet,out_sheet):
    COG_Mapper(in_sheet,outsheet)


# # Now that we mapped the COG ID's to the DE protein data set we now want to reduce the size of this data set 
# # and select specific columns that we care about. Those columns are: Protein Description, 
# # Averaged Log2Ratio, p-value, refseq_locus,refseq_protein_id, COG category, and the description of the COG catergory associated with the 
# # aligned protein.

workbook=openpyxl.Workbook()
workbook.save(filename="Proteomics_DE_COG_Desired_Columns.xlsx")

filtered_COG_prot_in = ["LowCa_vs_Ca","Ca_vs_Low_Fe","LowFe_vs_Fe","Fe_vs_Mix", "Mix_vs_Ni", 
"Ni_vs_Nd", "Nd_vs_W", "W_vs_Cu", "lowCa_vs_Cu"]

filtered_COG_prot_out = ["LowCa_vs_Ca","Ca_vs_Low_Fe","LowFe_vs_Fe","Fe_vs_Mix", "Mix_vs_Ni", 
"Ni_vs_Nd", "Nd_vs_W", "W_vs_Cu", "lowCa_vs_Cu"]

def cog_sorter(input_sheet,output_sheet):
    df = pd.read_excel("Filtered_DE_Prot_data.xlsx",sheet_name=input_sheet)
    desired_columns=["Description_x", "Averaged Log2Ratio", "p-value", "refseq_locus","refseq_protein_id","COG_category", "Description_y"]
    df_selected=df[desired_columns]
    path = "Proteomics_DE_COG_Desired_Columns.xlsx"
    with pd.ExcelWriter(path, mode='a', engine='openpyxl',if_sheet_exists="new") as writer:
        df_selected.to_excel(writer, sheet_name=output_sheet, index=False)

for input_sheet,output_sheet in zip(filtered_COG_prot_in,filtered_COG_prot_out):
    cog_sorter(input_sheet,output_sheet)


# #Now that we have a smaller data set with the desired columns, lets now split up the excel file into
# # mutliple sheets where each sheet will detail which proteins in each condition are mapped to a 
# # specific COG category we are interested in. In this case I am interested in any COG ID that is
# # associated with transporter function. If you look at the list called "Desired_COG" you will 
# # see which COG id's I am curious about. 

workbook=openpyxl.Workbook()
workbook.save(filename="Filtered_DEP_COG_Transporters.xlsx")

filtered_COG_prot= ["LowCa_vs_Ca","Ca_vs_Low_Fe","LowFe_vs_Fe","Fe_vs_Mix", "Mix_vs_Ni", 
 "Ni_vs_Nd", "Nd_vs_W", "W_vs_Cu", "lowCa_vs_Cu"]

sheet=["LowCa_vs_Ca","LowCa_vs_Ca1","LowCa_vs_Ca2","LowCa_vs_Ca3","LowCa_vs_Ca4",
"Ca_vs_Low_Fe", "Ca_vs_Low_Fe1","Ca_vs_Low_Fe2","Ca_vs_Low_Fe3","Ca_vs_Low_Fe4",
"LowFe_vs_Fe","LowFe_vs_Fe1","LowFe_vs_Fe2","LowFe_vs_Fe3","LowFe_vs_Fe4",
"Fe_vs_Mix","Fe_vs_Mix1","Fe_vs_Mix2","Fe_vs_Mix3","Fe_vs_Mix4",
"Mix_vs_Ni","Mix_vs_Ni1","Mix_vs_Ni2","Mix_vs_Ni3","Mix_vs_Ni4",
"Ni_vs_Nd","Ni_vs_Nd1","Ni_vs_Nd2","Ni_vs_Nd3","Ni_vs_Nd4",
"Nd_vs_W","Nd_vs_W1","Nd_vs_W2","Nd_vs_W3","Nd_vs_W4",
"W_vs_Cu","W_vs_Cu1","W_vs_Cu2","W_vs_Cu3","W_vs_Cu4",
"lowCa_vs_Cu","lowCa_vs_Cu1","lowCa_vs_Cu2","lowCa_vs_Cu3","lowCa_vs_Cu4"]

Desired_COG = ["H", "P", "U", "Q", "W"]

def COG_transporter_sorter(input_sheet):
    df=pd.read_excel("Proteomics_DE_COG_Desired_Columns.xlsx",sheet_name=input_sheet)
    path = "Filtered_DEP_COG_Transporters.xlsx"
    with pd.ExcelWriter(path,mode='a',engine='openpyxl', if_sheet_exists="new") as writer:
        for COG_id in Desired_COG:
            Filteted_Transporters = df[df['COG_category'] == COG_id]
            Filteted_Transporters.to_excel(writer, sheet_name=input_sheet, index=False)
for input_sheet in filtered_COG_prot:
     COG_transporter_sorter(input_sheet)

for sheetname in sheet:
    print(len(pd.read_excel("Filtered_DEP_COG_Transporters.xlsx",sheet_name=sheetname)))
